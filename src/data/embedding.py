from google.cloud import storage, bigquery
import google.generativeai as genai
import datetime
import os
import tempfile
import PyPDF2
from ..env import env
from typing import Tuple

from openpyxl import load_workbook

def get_clients() -> Tuple[storage.Client, bigquery.Client, genai]:
    """Get Google Cloud clients."""
    try:
        # Configurar Google AI
        genai.configure(api_key=env().google_cloud_api_key)
        
        return (
            storage.Client(),
            bigquery.Client(),
            genai
        )
    except Exception as e:
        raise RuntimeError(f"Failed to initialize Google Cloud clients: {e}")


def document_processing() -> None:
    """Process documents and generate embeddings."""
    storage_client, bq_client, genai_client = get_clients()
    
    # Usar variables de entorno
    project_id = env().project_id
    bucket_name = env().bucket_name
    dataset = env().dataset
    table = env().table
    
    bucket = storage_client.bucket(bucket_name)
    blobs = bucket.list_blobs()
    
    rows_to_insert = []

    for blob in blobs:
        if blob.name.endswith("/"):
            continue

        print(f"Procesando {blob.name}...")

        # Usar directorio temporal del sistema
        temp_dir = tempfile.gettempdir()
        local_path = os.path.join(temp_dir, blob.name.split('/')[-1])
        blob.download_to_filename(local_path)

        # Extraer texto según el tipo de archivo
        if local_path.lower().endswith('.pdf'):
            # Para PDFs usar PyPDF2
            with open(local_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                texto = ""
                for page in pdf_reader.pages:
                    texto += page.extract_text() + "\n"
        elif local_path.lower().endswith('.docx'):
            # Para DOCX usar python-docx
            try:
                from docx import Document
                doc = Document(local_path)
                texto = ""
                for paragraph in doc.paragraphs:
                    texto += paragraph.text + "\n"
            except ImportError:
                texto = f"Archivo DOCX: {blob.name} (necesita python-docx)"
        elif local_path.lower().endswith('.xlsx'):
            # Para XLSX usar openpyxl
            try:
                wb = load_workbook(local_path, read_only=True)
                texto = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell:
                                texto += str(cell) + " "
                        texto += "\n"
                wb.close()  # Cerrar el workbook explícitamente
            except ImportError:
                texto = f"Archivo XLSX: {blob.name} (necesita openpyxl)"
        else:
            # Para otros archivos, intentar leer como texto
            try:
                with open(local_path, 'r', encoding='utf-8') as file:
                    texto = file.read()
            except (UnicodeDecodeError, IOError) as e:
                print(f"Error reading file {blob.name}: {e}")
                texto = f"Archivo: {blob.name}"

        # Dividir en fragmentos
        chunks = [texto[i:i+1000] for i in range(0, len(texto), 1000)]

        # Limpiar archivo temporal después de procesarlo
        try:
            if os.path.exists(local_path):
                os.remove(local_path)
        except PermissionError:
            print(f"No se pudo eliminar el archivo temporal: {local_path}")
            # Continuar con el procesamiento

        for chunk in chunks:
            try:
                # Generar embedding usando Google AI
                embedding_result = genai_client.embed_content(
                    model="models/embedding-001",
                    content=chunk,
                    task_type="retrieval_document"
                )
                embedding = embedding_result['embedding']
            except Exception as e:
                print(f"Error generating embedding for {blob.name}: {e}")
                continue

            rows_to_insert.append({
                "document_id": blob.name,
                "fragment_text": chunk,
                "embedding": embedding,
                "created_at": datetime.datetime.now(datetime.timezone.utc).isoformat()
            })

    # Insertar en BigQuery
    errors = bq_client.insert_rows_json(f"{project_id}.{dataset}.{table}", rows_to_insert)
    if errors:
        print("Errores al insertar:", errors)
    else:
        print("Embeddings insertados correctamente en BigQuery.")


if __name__ == "__main__":
    document_processing()